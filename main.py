from openpyxl import load_workbook, Workbook
from json import load,dump
from datetime import datetime
from calendar import monthrange
from openpyxl.styles import PatternFill, Border, Side

with open("All_rooms/all_rooms.json", "r") as f:
    all_rooms_list = load(f)

with open("All_rooms/next_duty_list.json", "r") as f:
    next_duty_rooms = load(f)

current_time = datetime.now()
current_year = current_time.year
current_month = current_time.month
days = monthrange(current_year, current_month)
days = days[-1] + 1

Fill_cell_style = PatternFill(patternType='solid',
                       fgColor='808080')

def Room_merges():

    all_merged_rooms = {}
    for floors in all_rooms_list.keys():
    
        deficient_rooms_list = {}
        all_merged_rooms[floors] = {}

        for wing in all_rooms_list[floors].keys():

            room_list = []

            for current_room in all_rooms_list[floors][wing]:

                peoples_count = all_rooms_list[floors][wing][current_room]

                if peoples_count < 3:

                    if deficient_rooms_list != {}:

                        unused_deficent_rooms = {}

                        for deficient_rooms in deficient_rooms_list.copy().keys():

                            
                            if peoples_count + deficient_rooms_list[deficient_rooms]["Count_peoples"] < 3:

                                deficient_rooms_list[deficient_rooms]["Count_peoples"] += peoples_count
                                deficient_rooms_list[deficient_rooms]["Rooms"] += f', {current_room}'

                            elif peoples_count + deficient_rooms_list[deficient_rooms]["Count_peoples"] == 3:    
                                deficient_rooms_list[deficient_rooms]["Count_peoples"] += peoples_count
                                deficient_rooms_list[deficient_rooms]["Rooms"] += f', {current_room}'
                                # room_list.append(deficient_rooms_list[deficient_rooms]["Rooms"])
                                # del deficient_rooms_list[deficient_rooms]["Rooms"]

                            else:
                                deficient_rooms_list[current_room] = {"Count_peoples" : peoples_count, "Rooms" : f'{current_room}'}
                            # print(deficient_rooms_list)
                    else:

                        deficient_rooms_list[current_room] = {"Count_peoples" : peoples_count, "Rooms" : f'{current_room}'}
                        # print(deficient_rooms_list)

                else:
            
                    room_list.append(current_room)


            for deficient_rooms in deficient_rooms_list:

                room_list.append(deficient_rooms_list[deficient_rooms]["Rooms"])
            
            all_merged_rooms[floors][wing] = room_list
    return all_merged_rooms

def Create_tables():
    for floors in all_rooms_list_merged.keys():

        for wing in all_rooms_list_merged[floors].keys():

            wb = Workbook()
            sheet = wb.active

            start_index = 1
            for rooms in all_rooms_list_merged[floors][wing]:
                start_index += 1
                cell = sheet.cell(row=1, column=start_index) 
                cell.value = rooms

            for day in range(1,days):
                cell = sheet.cell(row=day+1, column=1) 
                cell.value = f'{day}.{current_month}'
            wb.save(filename=f'{floors}/{wing}.xlsx')
     
def Paint_cells():
    for floors in all_rooms_list_merged.keys():
        
        for wing in all_rooms_list_merged[floors].keys():
            path = f'{floors}/{wing}.xlsx'
            wb = load_workbook(path)
            sheet = wb.active
            row = sheet.max_row
            column = sheet.max_column

            

            for dates in range(2,row+1):
                next_duty = next_duty_rooms[floors][wing]
                print(next_duty)
                for room in range(2,column+1):
                    room_cell = sheet.cell(row=1, column=room)
                    room_value = room_cell.value

                    if room_value == next_duty:
                        duty_room = sheet.cell(row=dates, column=room)
                        duty_room.fill = Fill_cell_style
                        
                        next_duty = sheet.cell(row=1,column=int([sheet.cell(row=1, column= room + 1).column if room < column else sheet.cell(row=1, column= room + 2 - column).column][0])).value
                        print(next_duty)
                        # print(duty_room.column, next_duty.column)
                        break
                next_duty_rooms[floors][wing] = next_duty
                print(dates, wing, floors)

                    
            wb.save(filename=f'{floors}/{wing}.xlsx')      

    with open("All_rooms/next_duty_list.json", "w") as f:
                dump(next_duty_rooms,f,indent=4)       

def Table_style_create():
    pass
all_rooms_list_merged = Room_merges()
Create_tables()
Paint_cells()