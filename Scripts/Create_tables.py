from pathlib import Path
from datetime import datetime
from json import dump
from calendar import monthrange
from openpyxl import load_workbook, Workbook, styles

class Create_Tables:

    def __init__(self,merged_rooms, duty_rooms):

        current_time = datetime.now()
        self.current_year = current_time.year
        self.current_month = current_time.month
        days = monthrange(self.current_year, self.current_month)
        self.days = days[-1] + 1

        self.all_rooms_list_merged = merged_rooms
        self.next_duty_rooms = duty_rooms

        self.Fill_cell_style = styles.PatternFill(patternType='solid',
                        fgColor='808080')
        
        path = Path(f'Tables/{self.current_month}.{self.current_year}')
        path.mkdir(parents=True, exist_ok=True)

        self.createTables()
        self.fillTable()

    def createTables(self):
        for floors in self.all_rooms_list_merged.keys():

            for wing in self.all_rooms_list_merged[floors].keys():

                wb = Workbook()
                sheet = wb.active

                start_index = 1
                for rooms in self.all_rooms_list_merged[floors][wing]:
                    start_index += 1
                    cell = sheet.cell(row=1, column=start_index) 
                    cell.value = rooms

                for day in range(1,self.days):
                    cell = sheet.cell(row=day+1, column=1) 
                    cell.value = f'{day}.{self.current_month}'
                
                wb.save(filename=f'Tables/{self.current_month}.{self.current_year}/{floors}_{wing}.xlsx')
    
    def fillTable(self):
        for floors in self.all_rooms_list_merged.keys():
            
            for wing in self.all_rooms_list_merged[floors].keys():
                path = f'Tables/{self.current_month}.{self.current_year}/{floors}_{wing}.xlsx'
                wb = load_workbook(path)
                sheet = wb.active
                row = sheet.max_row
                column = sheet.max_column

                for dates in range(2,row+1):
                    next_duty = self.next_duty_rooms[floors][wing]
                    
                    for room in range(2,column+1):
                        room_cell = sheet.cell(row=1, column=room)
                        room_value = room_cell.value

                        if room_value == next_duty:
                            duty_room = sheet.cell(row=dates, column=room)
                            duty_room.fill = self.Fill_cell_style
                            
                            next_duty = sheet.cell(row=1,column=int([sheet.cell(row=1, column= room + 1).column if room < column else sheet.cell(row=1, column= room + 2 - column).column][0])).value
                            
                            # print(duty_room.column, next_duty.column)
                            break
                    self.next_duty_rooms[floors][wing] = next_duty
                    

                        
                wb.save(filename=f'Tables/{self.current_month}.{self.current_year}/{floors}_{wing}.xlsx')      

        with open("./data/next_duty_list.json", "w") as f:
                    dump(self.next_duty_rooms,f,indent=4)       
