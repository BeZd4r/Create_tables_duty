import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

class Add_cell_style:

    def __init__(self,current_time):

        self.current_year = current_time.year
        self.current_month = current_time.month
        
        self.border_style = Border(top = Side(border_style='thin', color='00000000'),    
                              right = Side(border_style='thin', color='00000000'), 
                              bottom = Side(border_style='thin', color='00000000'),
                              left = Side(border_style='thin', color='00000000'))

        pathToDir = f'Tables/{self.current_month}.{self.current_year}'
        self.tables = []

        for tables_name in os.listdir(pathToDir):
            table_name = f'{pathToDir}/{tables_name}'
            self.tables.append(table_name)
       
        self.decorateCell()

    def decorateCell(self):

        for table in self.tables:
            wb = load_workbook(table)
            ws = wb.active
            max_length = 0

            for i in range(1, ws.max_row+1): 
                
                for j in range(1, ws.max_column+1): 
                    cellt = ws.cell(row=i, column=j) 
                    cellt.border = self.border_style
                    cellt.alignment = Alignment(horizontal='center')           
            
            for col in ws.columns:
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            wb.save(table)